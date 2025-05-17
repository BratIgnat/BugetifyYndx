import os
import logging
from datetime import datetime, date
from dotenv import load_dotenv
from telegram import Update, InputFile
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters
from openpyxl import Workbook, load_workbook

from speechkit import recognize_ogg  # функция распознавания

# ─────────── Загрузка переменных окружения ───────────

load_dotenv()
BOT_TOKEN      = os.getenv("BOT_TOKEN")
YANDEX_API_KEY = os.getenv("YANDEX_API_KEY")

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN не задан в .env")
if not YANDEX_API_KEY:
    raise RuntimeError("YANDEX_API_KEY не задан в .env")

logging.basicConfig(level=logging.WARNING)  # Убираем лишние логи
logger = logging.getLogger(__name__)

# ─────────── Работа с Excel (индивидуальные файлы) ───────────

def get_user_excel_file(user_id):
    return f"expenses_{user_id}.xlsx"

def init_excel_file(excel_file):
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Сумма", "Категория", "Источник", "Позиции"])
        wb.save(excel_file)

def append_to_excel(user_id, amount, category, source="Голос", items="-"):
    excel_file = get_user_excel_file(user_id)
    init_excel_file(excel_file)
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        amount,
        category,
        source,
        items
    ])
    wb.save(excel_file)

def update_last_row(user_id, col_idx, new_value):
    excel_file = get_user_excel_file(user_id)
    init_excel_file(excel_file)
    wb = load_workbook(excel_file)
    ws = wb.active
    if ws.max_row < 2:
        raise Exception("В вашей таблице пока нет ни одной записи")
    ws.cell(row=ws.max_row, column=col_idx).value = new_value
    wb.save(excel_file)

# ─────────── Telegram-команды ───────────

async def add_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    try:
        if len(context.args) < 2:
            await update.message.reply_text("Формат: /add сумма категория")
            return
        amount = float(context.args[0].replace(",", "."))
        category = " ".join(context.args[1:])
        append_to_excel(user_id, amount, category, source="Команда")
        await update.message.reply_text(f"✅ Добавлено: {amount} — {category}")
    except ValueError:
        await update.message.reply_text("❌ Ошибка: сумма должна быть числом.")
    except Exception as e:
        logger.exception(e)
        await update.message.reply_text(f"❌ Ошибка: {e}")

async def edit_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    try:
        if len(context.args) != 1:
            await update.message.reply_text("Формат: /edit_amount новая_сумма")
            return
        new_amount = float(context.args[0].replace(",", "."))
        update_last_row(user_id, col_idx=2, new_value=new_amount)
        await update.message.reply_text(f"✅ Сумма обновлена на {new_amount}")
    except Exception as e:
        logger.exception(e)
        await update.message.reply_text(f"❌ Ошибка: {e}")

async def edit_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    try:
        if not context.args:
            await update.message.reply_text("Формат: /edit_category новая_категория")
            return
        new_category = " ".join(context.args)
        update_last_row(user_id, col_idx=3, new_value=new_category)
        await update.message.reply_text(f"✅ Категория обновлена на: {new_category}")
    except Exception as e:
        logger.exception(e)
        await update.message.reply_text(f"❌ Ошибка: {e}")

async def send_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        await update.message.reply_text("У вас ещё нет расходов! Добавьте запись — и файл появится.")
        return
    try:
        with open(excel_file, "rb") as f:
            await update.message.reply_document(document=InputFile(f, filename=f"expenses_{user_id}.xlsx"))
    except Exception as e:
        await update.message.reply_text(f"Ошибка при отправке файла: {e}")

async def last_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    excel_file = get_user_excel_file(user_id)
    if not os.path.exists(excel_file):
        await update.message.reply_text("У вас ещё нет расходов за сегодня!")
        return
    wb = load_workbook(excel_file)
    ws = wb.active
    today_str = date.today().strftime("%Y-%m-%d")
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0].startswith(today_str):
            records.append(row)
    if not records:
        await update.message.reply_text("Сегодня ещё нет записей.")
        return
    msg = "\n".join(
        [f"{r[0]} — {r[1]} ₽ — {r[2]} ({r[3]})" for r in records]
    )
    await update.message.reply_text(f"Записи за сегодня:\n{msg}")

# ─────────── Голосовые сообщения ───────────

async def voice_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    try:
        voice = update.message.voice
        file = await context.bot.get_file(voice.file_id)
        file_path = f"voice_{update.message.message_id}_{user_id}.ogg"
        await file.download_to_drive(file_path)

        text = recognize_ogg(file_path)
        logger.info(f"Распознан текст: {text!r}")

        # Парсим сумму и категорию
        words = text.strip().split()
        amount = None
        category = None
        for i, w in enumerate(words):
            try:
                amount = float(w.replace(",", "."))
                category = " ".join(words[i+1:]) or "Без категории"
                break
            except ValueError:
                continue
        if amount is None:
            raise ValueError("Не удалось найти сумму в распознанном тексте.")

        append_to_excel(user_id, amount, category, source="Голос")
        await update.message.reply_text(f"✅ Добавлено: {amount} — {category}")
    except Exception as e:
        logger.exception(e)
        await update.message.reply_text(f"❌ Ошибка при обработке голоса: {e}")
    finally:
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)

# ─────────── Запуск бота ───────────

def main():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("add", add_command))
    app.add_handler(CommandHandler("edit_amount", edit_amount))
    app.add_handler(CommandHandler("edit_category", edit_category))
    app.add_handler(CommandHandler("get_excel", send_excel))
    app.add_handler(CommandHandler("last", last_command))
    app.add_handler(MessageHandler(filters.VOICE, voice_handler))
    print("Бот запущен! Ожидаю сообщений...")
    app.run_polling()

if __name__ == "__main__":
    main()
