# excel_writer.py
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

def write_to_excel(amount, category, source="Голос", positions=None):
    filename = "expenses.xlsx"

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Сумма", "Категория", "Источник", "Позиции"])
        wb.save(filename)

    wb = load_workbook(filename)
    ws = wb.active

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    positions_text = "\n".join(positions) if positions else "-"
    ws.append([now, amount, category, source, positions_text])
    wb.save(filename)