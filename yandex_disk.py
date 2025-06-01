import os
import requests
import io
import openpyxl
import re
import json
from datetime import datetime
import pytz

TOKENS_DIR = "tokens"
SETTINGS_DIR = "user_settings"
if not os.path.exists(TOKENS_DIR):
    os.makedirs(TOKENS_DIR)
if not os.path.exists(SETTINGS_DIR):
    os.makedirs(SETTINGS_DIR)

class ExpenseParseError(Exception):
    """Кастомное исключение для ошибок парсинга и валидации расходов."""
    pass

def get_auth_link(user_id):
    client_id = os.getenv("YANDEX_CLIENT_ID")
    return (
        f"https://oauth.yandex.ru/authorize?response_type=code&client_id={client_id}&scope=cloud_api:disk.app_folder"
    )

def set_auth_code(user_id, code):
    client_id = os.getenv("YANDEX_CLIENT_ID")
    client_secret = os.getenv("YANDEX_CLIENT_SECRET")
    url = "https://oauth.yandex.ru/token"
    data = {
        "grant_type": "authorization_code",
        "code": code,
        "client_id": client_id,
        "client_secret": client_secret,
    }
    r = requests.post(url, data=data)
    if r.status_code == 200:
        token = r.json().get("access_token")
        with open(f"{TOKENS_DIR}/{user_id}.token", "w") as f:
            f.write(token)
        return True
    return False

def is_user_authenticated(user_id):
    return os.path.exists(f"{TOKENS_DIR}/{user_id}.token")

def get_user_token(user_id):
    try:
        with open(f"{TOKENS_DIR}/{user_id}.token", "r") as f:
            return f.read().strip()
    except Exception:
        return None

def parse_expense(text):
    """
    Возвращает (amount, category) если оба элемента есть и валидны, иначе (None, None)
    """
    original_text = (text or "").lower().strip()
    if not original_text:
        return None, None

    # Ищем сумму
    pattern = r"(\d+[.,]?\d*)\s*(руб(ль|лей|ля|ли|лем|лям|лями)?|р)?(\s*\d{1,2}\s*(коп(ейка|ейки|еек|ейку|ейкой|ейками)?|к)?)?"
    matches = list(re.finditer(pattern, original_text, flags=re.IGNORECASE))
    if not matches:
        return None, None

    for match in matches:
        if match:
            rub = match.group(1)
            kop = None
            if match.group(4):
                kop_match = re.search(r"\d{1,2}", match.group(4))
                if kop_match:
                    kop = kop_match.group(0)
            if kop:
                amount = float(rub) + float(kop)/100
            else:
                amount = float(rub)
            text_wo_amount = original_text.replace(match.group(0), "")
            category = re.sub(r"\s+", " ", text_wo_amount).strip()
            # Жёсткая проверка: и сумма, и категория должны быть!
            if not category or not amount or category == "":
                return None, None
            return amount, category

    return None, None

def save_timezone(user_id, timezone_str):
    path = f"{SETTINGS_DIR}/{user_id}.json"
    with open(path, "w") as f:
        json.dump({"timezone": timezone_str}, f)

def get_timezone(user_id):
    path = f"{SETTINGS_DIR}/{user_id}.json"
    try:
        with open(path, "r") as f:
            data = json.load(f)
            return data.get("timezone", "UTC")
    except:
        return "UTC"

def save_to_yadisk(user_id, text, message_date=None):
    token = get_user_token(user_id)
    if not token:
        raise Exception("User not authenticated")
    file_name = f"{user_id}.xlsx"
    remote_path = f"app:/{file_name}"

    amount, category = parse_expense(text)
    if amount is None or category is None:
        raise ExpenseParseError("Сообщение должно содержать сумму и категорию. Пример: '100 рублей такси', '150 кефир'.")

    headers = {"Authorization": f"OAuth {token}"}
    download_url = "https://cloud-api.yandex.net/v1/disk/resources/download"
    params_download = {"path": remote_path}
    r2 = requests.get(download_url, params=params_download, headers=headers)

    if r2.status_code == 200 and "href" in r2.json():
        download_href = r2.json()["href"]
        file_content = requests.get(download_href).content
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["#", "Сумма", "Категория", "Дата/Время"])

    if sheet.max_column < 4 or (sheet.cell(row=1, column=4).value or "").strip().lower() not in ["дата/время", "дата и время"]:
        sheet.cell(row=1, column=4).value = "Дата/Время"

    idx = sheet.max_row

    user_tz_str = get_timezone(user_id)
    try:
        user_tz = pytz.timezone(user_tz_str)
    except Exception:
        user_tz = pytz.UTC

    if message_date:
        dt_utc = datetime.fromtimestamp(message_date, tz=pytz.UTC)
        dt = dt_utc.astimezone(user_tz)
    else:
        dt = datetime.now(user_tz)
    dt_str = dt.strftime("%Y-%m-%d %H:%M:%S")

    sheet.append([idx, amount, category, dt_str])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    upload_url = "https://cloud-api.yandex.net/v1/disk/resources/upload"
    params = {"path": remote_path, "overwrite": "true"}
    r = requests.get(upload_url, params=params, headers=headers)
    href = r.json().get("href")
    if not href:
        raise Exception("Не удалось получить ссылку для загрузки файла на Яндекс.Диск.")
    requests.put(href, data=output.getvalue())
